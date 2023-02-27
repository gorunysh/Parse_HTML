import requests
from requests.auth import HTTPBasicAuth
from openpyxl import load_workbook
from copy import copy
import json
import subprocess


from openpyxl.cell.text import InlineFont
from openpyxl.cell.text import RichText

# <th> = += 1 столбоец
# <tr> = следующия строка
# <td> = следующая ячейка

class Adventure_time():

    def registration_and_load_HTML(self, url, login, password) -> bool:
        try:
            response = requests.post(url, auth=HTTPBasicAuth(login, password))
            status_code = response.status_code
        except:
            status_code = 404

        if status_code == 200:
            self.txt_HTML = response.text
            result = True
        elif status_code == 401:
            print('Не верный логин или пароль')
            result = False
        elif status_code != 404:
            print('Ошибка соединения, статус код =', status_code)
            result = False
        else:
            print('Ошибка подключения к wiki =', status_code)
            result = False
        return result

    def registration_and_load_HTML_for_QT(self, url, login, password) -> int:
        try:
            response = requests.post(url, auth=HTTPBasicAuth(login, password))
            status_code = response.status_code

            if status_code == 200:
                self.txt_HTML = response.text
        except:
            status_code = 404

        return status_code

    def parse_HTML(self, number_table: str):
        txt = self.txt_HTML
        number_table = int(number_table)

        # Выбор нужной таблицы + обрезка лишнего
        for i in range(number_table - 1):
            txt = txt[txt.find('</table>') + 7:]

        # обрезка текста HTML
        if not number_table - 1:
            txt = txt[txt.find('<table'):txt.find('</table>')]
        else:
            txt = txt[:txt.find('</table>') + 8]
        txt = txt[txt.find('<th'):]

        # деление по строкам
        txt_list = list(txt.split('<tr>'))

        # деление по ячейкам
        for nomber, i in enumerate(txt_list):
            if '</th><th' in i:
                txt_list[nomber] = i.split('</th><th')
            elif '</td><td' in i:
                txt_list[nomber] = i.split('</td><td')

        result = list()
        # <strong> - жирный текст
        # <ins> - подчеркивание
        # <i> - Курсив

        # чистка текста
        for number_i, i in enumerate(txt_list):
            result.append([])
            for number_j, j in enumerate(i):
                result[number_i].append({'txt': '', 'http': '', 'font': False})

                # чистка от nbsp в пустых строках
                if '''px">&nbsp;''' in j:
                    j = ' '

                # работа с ссылками
                elif 'href' in j:
                    result[number_i][number_j]['http'] = j[j.find('href="')+6:]
                    result[number_i][number_j]['http'] = result[number_i][number_j]['http'][:result[number_i][number_j]['http'].find('"')]
                else:
                    j = j[j.find('>')+1:]

                # работа и исключениями, из HTML
                except_list = ['<strong>', '<ins>', '<i>']
                for except_ in except_list:
                    if except_ in j:
                        result[number_i][number_j]['font'] = True

                # табуляция и переносы текста внутри ячейки
                while '</p>' in j:
                    j = j[:j.find('</p>')] + f'\n'*2 + j[j.find('</p>')+4:]
                while '<br/>' in j:
                    j = j[:j.find('<br/>')] + f'\n' + j[j.find('<br/>')+5:]
                counter = 1
                while '<li>' in j:
                    j = j[:j.find('<li>')] + f'\n {counter}. ' + j[j.find('<li>')+4:]
                    counter +=1

                # окончательная чистка текста в ячейке от артефактов
                if j.find('<') > j.find('>'):
                    j = j[j.find('>')+1:]
                while '<' in j and '>' in j:
                    j = j[:j.find('<')] + j[j.find('>')+1:]
                while 'nbsp' in j:
                    j = j[:j.find('nbsp')-1] + j[j.find('nbsp')+5:]

                result[number_i][number_j]['txt'] = j

        len_norm = len(result[1])

        # компенсирует обьединенные ячейки в первых столбцах
        for i in range(3, len(result)):
            try:
                pobe = int(result[i][0])
            except:
                while len(result[i]) < len_norm:
                    result[i].insert(0, {'txt': ' ', 'http': '', 'font': False})

        self.result = result
        return result

    def save_xslx_v4_for_QT(self, pattern_name: str, save_name: str):

        try:
            pattern_wb = load_workbook(pattern_name)
            new_wb = copy(pattern_wb)
            new_ws = new_wb.active

            for number_row, row in enumerate(self.result):
                for number_cell, cell in enumerate(row):

                    new_ws.cell(number_row + 1, number_cell + 1).value = cell['txt']

            message_text = (
                    f'Таблица сформированна: строк = {len(self.result) - 1}, столбцов = {len(self.result[1]) - 1}')
        except Exception as e:
            message_text = ('не удалось загрузить шаблон, может быть не правильный формат, должен быть: .xlsx')
            print(e)

        message_text +='''
'''

        try:
            new_wb.save(save_name)
            message_text += (f'Файл сохранен: {save_name}')
        except:
            message_text += 'не удалось сохранить файл, проверти формат для сохранения (должен быть .xlsx или .xls)'

        return message_text

    def save_xslx_fontHTML_for_QT(self, save_name: str):
        new_wb = load_workbook(save_name)
        new_ws = new_wb.active

        red = InlineFont(color='FF000000')

        new_ws['A1'] = RichText(rPr=red, t='When the color ')

        new_wb.save(save_name)


def input_config(data: dict, config_name: str) -> dict:
    notifications = {
        'login': 'Введите login ', 'password': 'Введите пароль ', 'url': 'Добавьте адрес Xwiki ',
        'pattern_name': 'Введите имя шаблона ', 'save_name': 'Введите имя файла для сохранения ',
        'number_table': 'Введите номер таблицы для использования '
    }
    questions = {
        'save': 'Желаете сохранить для использования в будущем? Введите: да/нет'
    }

    # проверка заполнения все полей
    edit_json_bool = False
    for kye, value in notifications.items():
        if data[kye] == '':
            data[kye] = input(value)
            edit_json_bool = True

    # запись в джейсон новых данных по желанию
    if edit_json_bool:
        if need_save_json():
            with open(config_name, 'w') as f:
                json.dump(data, f)

    return data

def need_save_json() -> bool:
    result = None
    input_ = input('Если надо сохранить введенные данные напишите (y/n) ')

    if input_ == 'y':
        print('Сохранил')
        result = True
    else:
        result = False

    return result

def open_json(config_name: str) -> dict:
    # чтение джейсона и работа с ним
    try:
        with open(config_name, 'r') as f:
            data_json = json.load(f)
            data = input_config(data_json, config_name)

    except:
        print('Не вижу файл с настройками:', config_name)

    # запись данных из джейсона
    try:
        result = {
            'login': data['login'],
            'password': data['password'],
            'url': data['url'],
            'pattern_name': data['pattern_name'],
            'save_name': data['save_name'],
            'number_table': data['number_table'],
        }
    except:
        print(f'Файл {config_name} или поврежден или заполнен не правильно, пример:'
              '''
  {
  "login": "логин",
  "password": "пароль",
  "url": "http://wiki.gt/bin/view/07.%20Развитие/Поставщики",
  "pattern_name": "pattern.xlsx",
  "save_name": "test_save.xlsx",
  "number_table": 2
  }
''')
    return result

def save_json_for_QT(data: dict, name_save: str):
    with open(name_save, 'w') as f:
        json.dump(data, f)

def open_json_for_QT(config_name: str) -> dict:

    # чтение джейсона и работа с ним
    with open(config_name, 'r') as f:
        result = json.load(f)

    notifications = {
        'login': 'Введите login Xwiki', 'password': 'Введите пароль Xwiki', 'url': 'Добавьте адрес Xwiki',
        'pattern_name': 'Введите имя шаблона', 'save_name': 'Введите имя файла для сохранения',
        'number_table': 'Введите номер таблицы в Xwiki'
    }

    for key, value in result.items():
        if value == '':
            result[key] = notifications[key]

    return result

def test_print_list(list_:list):
    for i in list_:
        print('='*200)
        if type(i) == list:
            for j in i:
                if type(j) == dict:
                    for j1 in j.values():
                        print(j1)
                else:
                    print(j)
        else:
            print('str======', i)

def open_Excel(file_name: str):
    try:
        subprocess.call(['soffice', file_name])
    except:
        print('Open excel')
        try:
            subprocess.call(['excel', file_name])
        except Exception as e:
            print(e)
